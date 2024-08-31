import openpyxl
import logging
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Tuple, Optional

class SpreadsheetManager:
    def __init__(self, file_path: str):
        self.file_path = file_path
        try:
            self.workbook = openpyxl.load_workbook(file_path)
        except (FileNotFoundError, InvalidFileException) as e:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(file_path)
            logging.info(f"Created a new workbook due to: {e}")

    def read_data(self, cell_range: str, sheet_name: Optional[str] = None) -> List[List[Optional[str]]]:
        try:
            if sheet_name:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.active
            return [[cell.value for cell in row] for row in sheet[cell_range]]
        except Exception as e:
            logging.error(f"Error reading data from {cell_range}: {e}")
            return []

    def write_data(self, start_cell: Tuple[int, int], data: List[List[Optional[str]]], sheet_name: Optional[str] = None) -> None:
        try:
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    self.workbook.create_sheet(title=sheet_name)
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.active
            for i, row in enumerate(data):
                for j, value in enumerate(row):
                    # Ensure that only valid data types are written to the spreadsheet
                    if isinstance(value, (int, float, type(None))):
                        sheet.cell(row=start_cell[0] + i, column=start_cell[1] + j, value=value)
                    else:
                        sheet.cell(row=start_cell[0] + i, column=start_cell[1] + j, value=str(value))
            self.workbook.save(self.file_path)
            logging.info(f"Data written successfully to {start_cell} in sheet '{sheet_name or self.workbook.active.title}'.")
        except Exception as e:
            logging.error(f"Error writing data to {start_cell} in sheet '{sheet_name or self.workbook.active.title}': {e}")

    def add_sheet(self, sheet_name: str) -> None:
        try:
            self.workbook.create_sheet(title=sheet_name)
            self.workbook.save(self.file_path)
        except Exception as e:
            print(f"Error adding sheet {sheet_name}: {e}")
